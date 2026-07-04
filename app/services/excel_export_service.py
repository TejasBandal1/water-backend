from __future__ import annotations

from datetime import date, datetime, time, timezone
from io import BytesIO
from typing import Literal
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from app.models.invoice import Invoice
from app.models.invoice_item import InvoiceItem
from app.models.payment import Payment
from app.models.trip import Trip
from app.models.trip_container import TripContainer

ExportReportType = Literal["invoices", "billing", "payments", "full"]

ALLOWED_STATUSES = {"draft", "pending", "partial", "overdue", "paid", "cancelled"}
IST = ZoneInfo("Asia/Kolkata")


def _to_money(value: float | int | None) -> float:
    return round(float(value or 0), 2)


def _to_ist_datetime(value: datetime | None) -> datetime | None:
    """Convert a naive UTC database timestamp to an Excel-safe IST datetime."""
    if value is None:
        return None

    utc_value = value.replace(tzinfo=timezone.utc) if value.tzinfo is None else value
    return utc_value.astimezone(IST).replace(tzinfo=None)


def _client_sort_key(name: str | None, client_id: int | None) -> tuple[str, int]:
    return ((name or "").strip().casefold(), int(client_id or 0))


def _resolved_status(invoice: Invoice) -> str:
    if (
        invoice.status == "pending"
        and invoice.due_date
        and datetime.utcnow() > invoice.due_date
    ):
        return "overdue"

    return (invoice.status or "").strip().lower()


def _to_datetime_range(
    from_date: date | None,
    to_date: date | None
) -> tuple[datetime | None, datetime | None]:
    # Treat UI dates as IST calendar dates, then convert to UTC for DB filters.
    # The DB stores UTC timestamps, so this prevents "blank export" for local-day filters.
    start_dt = None
    end_dt = None

    if from_date:
        start_local = datetime.combine(from_date, time.min).replace(tzinfo=IST)
        start_dt = start_local.astimezone(timezone.utc).replace(tzinfo=None)

    if to_date:
        end_local = datetime.combine(to_date, time.max).replace(tzinfo=IST)
        end_dt = end_local.astimezone(timezone.utc).replace(tzinfo=None)

    return start_dt, end_dt


def _write_sheet(ws, headers: list[str], rows: list[list]):
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    money_keywords = ("amount", "price", "line total", "billed total", "invoice total")
    integer_keywords = ("count", "quantity", "qty", "jars", "delivered", "returned", "trips")

    for column_index, header in enumerate(headers, start=1):
        normalized_header = header.casefold()
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=column_index)
            if any(keyword in normalized_header for keyword in money_keywords):
                cell.number_format = '₹#,##0.00'
            elif any(keyword in normalized_header for keyword in integer_keywords):
                cell.number_format = '#,##0'
            elif "date" in normalized_header or " at" in normalized_header:
                cell.number_format = 'yyyy-mm-dd hh:mm'

            if isinstance(cell.value, str) and len(cell.value) > 48:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws.row_dimensions[row_index].height = max(
                    ws.row_dimensions[row_index].height or 15,
                    30
                )

    if headers == ["Metric", "Value"]:
        for row_index in range(2, ws.max_row + 1):
            metric = str(ws.cell(row=row_index, column=1).value or "").casefold()
            value_cell = ws.cell(row=row_index, column=2)
            if "amount" in metric:
                value_cell.number_format = '₹#,##0.00'
            elif any(word in metric for word in ("count", "quantity", "delivered", "returned")):
                value_cell.number_format = '#,##0'
            elif "generated at" in metric:
                value_cell.number_format = 'yyyy-mm-dd hh:mm'

    subtotal_fill = PatternFill("solid", fgColor="E2E8F0")
    grand_total_fill = PatternFill("solid", fgColor="DBEAFE")
    for row in ws.iter_rows(min_row=2):
        row_values = [cell.value for cell in row]
        is_grand_total = any(value == "ALL CLIENTS" for value in row_values)
        is_client_total = any(
            isinstance(value, str) and value.endswith(" TOTAL")
            for value in row_values
        )
        if is_grand_total or is_client_total:
            for cell in row:
                cell.font = Font(bold=True, color="0F172A")
                cell.fill = grand_total_fill if is_grand_total else subtotal_fill

        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("CHECK"):
                cell.font = Font(bold=True, color="991B1B")
                cell.fill = PatternFill("solid", fgColor="FEE2E2")

    for col_cells in ws.columns:
        max_len = 0
        for cell in col_cells:
            text = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(text))
        column_letter = col_cells[0].column_letter
        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 48)


def _build_invoices_data(
    db: Session,
    start_dt: datetime | None,
    end_dt: datetime | None,
    client_id: int | None,
    driver_id: int | None,
    status: str | None,
    include_cancelled: bool
) -> tuple[list[Invoice], list[list], list[list]]:
    invoices_query = (
        db.query(Invoice)
        .options(
            joinedload(Invoice.client),
            joinedload(Invoice.trips).joinedload(Trip.driver)
        )
    )

    if start_dt:
        invoices_query = invoices_query.filter(Invoice.created_at >= start_dt)
    if end_dt:
        invoices_query = invoices_query.filter(Invoice.created_at <= end_dt)
    if client_id:
        invoices_query = invoices_query.filter(Invoice.client_id == client_id)
    if driver_id:
        invoices_query = invoices_query.filter(Invoice.trips.any(Trip.driver_id == driver_id))
    if not include_cancelled:
        invoices_query = invoices_query.filter(Invoice.status != "cancelled")
    if status and status != "overdue":
        invoices_query = invoices_query.filter(Invoice.status == status)

    invoices = (
        invoices_query
        .order_by(Invoice.created_at.desc(), Invoice.id.desc())
        .all()
    )

    if status:
        invoices = [inv for inv in invoices if _resolved_status(inv) == status]

    invoice_rows: list[list] = []
    for invoice in invoices:
        driver_names = sorted({
            (trip.driver.name or "").strip()
            for trip in (invoice.trips or [])
            if trip.driver and (trip.driver.name or "").strip()
        })
        total_amount = _to_money(invoice.total_amount)
        amount_paid = _to_money(invoice.amount_paid)
        outstanding = _to_money(max(total_amount - amount_paid, 0))

        invoice_rows.append([
            invoice.id,
            invoice.client_id,
            invoice.client.name if invoice.client else None,
            ", ".join(driver_names) if driver_names else None,
            _resolved_status(invoice),
            _to_ist_datetime(invoice.created_at),
            _to_ist_datetime(invoice.due_date),
            _to_ist_datetime(invoice.confirmed_at),
            total_amount,
            amount_paid,
            outstanding,
            len(invoice.trips or [])
        ])

    invoice_rows.sort(
        key=lambda row: (
            *_client_sort_key(row[2], row[1]),
            row[5] or datetime.min,
            row[0]
        )
    )

    invoice_ids = [invoice.id for invoice in invoices]
    invoice_item_rows: list[list] = []

    if invoice_ids:
        item_query = (
            db.query(InvoiceItem, Invoice)
            .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
            .options(
                joinedload(InvoiceItem.container),
                joinedload(Invoice.client)
            )
            .filter(InvoiceItem.invoice_id.in_(invoice_ids))
            .order_by(InvoiceItem.invoice_id.asc(), InvoiceItem.id.asc())
        )

        for item, invoice in item_query.all():
            invoice_item_rows.append([
                item.invoice_id,
                invoice.client_id if invoice else None,
                invoice.client.name if invoice and invoice.client else None,
                item.container_id,
                item.container.name if item.container else None,
                int(item.quantity or 0),
                _to_money(item.price_snapshot),
                _to_money(item.total)
            ])

    invoice_date_by_id = {row[0]: row[5] for row in invoice_rows}
    invoice_item_rows.sort(
        key=lambda row: (
            *_client_sort_key(row[2], row[1]),
            invoice_date_by_id.get(row[0]) or datetime.min,
            row[0],
            (row[4] or "").strip().casefold(),
            row[3] or 0
        )
    )

    return invoices, invoice_rows, invoice_item_rows


def _build_trips_data(
    db: Session,
    start_dt: datetime | None,
    end_dt: datetime | None,
    client_id: int | None,
    driver_id: int | None,
    status: str | None,
    include_cancelled: bool
) -> list[list]:
    trips_query = (
        db.query(Trip)
        .options(
            joinedload(Trip.client),
            joinedload(Trip.driver),
            joinedload(Trip.invoice)
        )
    )

    if start_dt:
        trips_query = trips_query.filter(Trip.created_at >= start_dt)
    if end_dt:
        trips_query = trips_query.filter(Trip.created_at <= end_dt)
    if client_id:
        trips_query = trips_query.filter(Trip.client_id == client_id)
    if driver_id:
        trips_query = trips_query.filter(Trip.driver_id == driver_id)

    trips = trips_query.order_by(Trip.created_at.desc(), Trip.id.desc()).all()

    trip_ids = [trip.id for trip in trips]
    trip_container_map: dict[int, list[TripContainer]] = {}

    if trip_ids:
        trip_containers = (
            db.query(TripContainer)
            .options(joinedload(TripContainer.container))
            .filter(TripContainer.trip_id.in_(trip_ids))
            .all()
        )

        for trip_container in trip_containers:
            trip_container_map.setdefault(trip_container.trip_id, []).append(trip_container)

    trip_rows: list[list] = []

    for trip in trips:
        invoice = trip.invoice
        invoice_status = _resolved_status(invoice) if invoice else None

        if status:
            if not invoice_status or invoice_status != status:
                continue

        if not include_cancelled and invoice_status == "cancelled":
            continue

        containers = trip_container_map.get(trip.id, [])
        delivered_total = sum(int(tc.delivered_qty or 0) for tc in containers)
        returned_total = sum(int(tc.returned_qty or 0) for tc in containers)

        container_breakdown = "; ".join(
            f"{tc.container.name if tc.container else f'Container {tc.container_id}'} "
            f"(D:{int(tc.delivered_qty or 0)} R:{int(tc.returned_qty or 0)})"
            for tc in sorted(
                containers,
                key=lambda x: (x.container.name.lower() if x.container and x.container.name else "")
            )
        )

        trip_rows.append([
            trip.id,
            _to_ist_datetime(trip.created_at),
            trip.client_id,
            trip.client.name if trip.client else None,
            trip.driver_id,
            trip.driver.name if trip.driver else None,
            trip.invoice_id,
            invoice_status,
            delivered_total,
            returned_total,
            container_breakdown
        ])

    trip_rows.sort(
        key=lambda row: (
            *_client_sort_key(row[3], row[2]),
            row[1] or datetime.min,
            row[0]
        )
    )

    return trip_rows


def _build_payments_data(
    db: Session,
    start_dt: datetime | None,
    end_dt: datetime | None,
    client_id: int | None,
    driver_id: int | None,
    status: str | None,
    payment_method: str | None,
    include_cancelled: bool
) -> list[list]:
    payments_query = (
        db.query(Payment)
        .options(
            joinedload(Payment.invoice)
            .joinedload(Invoice.client),
            joinedload(Payment.invoice)
            .joinedload(Invoice.trips)
            .joinedload(Trip.driver)
        )
    )

    if start_dt:
        payments_query = payments_query.filter(Payment.created_at >= start_dt)
    if end_dt:
        payments_query = payments_query.filter(Payment.created_at <= end_dt)
    if payment_method:
        payments_query = payments_query.filter(Payment.method == payment_method)
    if client_id:
        payments_query = payments_query.join(Invoice, Payment.invoice_id == Invoice.id).filter(
            Invoice.client_id == client_id
        )

    payments = (
        payments_query
        .order_by(Payment.created_at.desc(), Payment.id.desc())
        .all()
    )

    payment_rows: list[list] = []

    for payment in payments:
        invoice = payment.invoice
        invoice_status = _resolved_status(invoice) if invoice else None

        if status:
            if not invoice_status or invoice_status != status:
                continue

        if not include_cancelled and invoice_status == "cancelled":
            continue

        if driver_id:
            invoice_driver_ids = {
                trip.driver_id
                for trip in (invoice.trips or [])
            } if invoice else set()
            if driver_id not in invoice_driver_ids:
                continue

        payment_rows.append([
            payment.id,
            _to_ist_datetime(payment.created_at),
            payment.invoice_id,
            invoice.client_id if invoice else None,
            invoice.client.name if invoice and invoice.client else None,
            invoice_status,
            payment.method,
            _to_money(payment.amount),
            _to_money(payment.cash_amount),
            _to_money(payment.upi_amount),
            payment.upi_account
        ])

    payment_rows.sort(
        key=lambda row: (
            *_client_sort_key(row[4], row[3]),
            row[1] or datetime.min,
            row[0]
        )
    )

    return payment_rows


def _build_client_views(
    invoice_rows: list[list],
    invoice_item_rows: list[list],
    trip_rows: list[list],
    payment_rows: list[list]
) -> tuple[list[list], list[list], dict]:
    """Build user-facing client-grouped rows without changing source records."""
    clients: dict[int, dict] = {}

    def get_client(client_id: int | None, client_name: str | None) -> dict:
        key = int(client_id or 0)
        if key not in clients:
            clients[key] = {
                "client_id": client_id,
                "client_name": (client_name or "Unknown Client").strip(),
                "invoice_count": 0,
                "billed_quantity": 0,
                "billed_amount": 0.0,
                "paid_amount": 0.0,
                "outstanding_amount": 0.0,
                "payments_in_period": 0.0,
                "trip_count": 0,
                "delivered_in_period": 0,
                "returned_in_period": 0,
                "uninvoiced_delivered": 0,
                "reconciliation": "N/A"
            }
        elif client_name and clients[key]["client_name"] == "Unknown Client":
            clients[key]["client_name"] = client_name.strip()
        return clients[key]

    invoice_by_id = {row[0]: row for row in invoice_rows}
    items_by_invoice: dict[int, list[list]] = {}
    item_totals_by_invoice: dict[int, float] = {}
    invoice_line_valid: dict[int, bool] = {}

    for row in invoice_rows:
        client = get_client(row[1], row[2])
        client["invoice_count"] += 1
        client["billed_amount"] = _to_money(client["billed_amount"] + row[8])
        client["paid_amount"] = _to_money(client["paid_amount"] + row[9])
        client["outstanding_amount"] = _to_money(
            client["outstanding_amount"] + row[10]
        )

    for row in invoice_item_rows:
        client = get_client(row[1], row[2])
        client["billed_quantity"] += int(row[5] or 0)
        items_by_invoice.setdefault(row[0], []).append(row)
        item_totals_by_invoice[row[0]] = _to_money(
            item_totals_by_invoice.get(row[0], 0) + row[7]
        )
        line_is_valid = _to_money((row[5] or 0) * (row[6] or 0)) == _to_money(row[7])
        invoice_line_valid[row[0]] = invoice_line_valid.get(row[0], True) and line_is_valid

    for row in trip_rows:
        client = get_client(row[2], row[3])
        client["trip_count"] += 1
        client["delivered_in_period"] += int(row[8] or 0)
        client["returned_in_period"] += int(row[9] or 0)
        if row[6] is None:
            client["uninvoiced_delivered"] += int(row[8] or 0)

    for row in payment_rows:
        client = get_client(row[3], row[4])
        client["payments_in_period"] = _to_money(
            client["payments_in_period"] + row[7]
        )

    invoice_reconciliation: dict[int, str] = {}
    for invoice_id, invoice_row in invoice_by_id.items():
        has_items = bool(items_by_invoice.get(invoice_id))
        totals_match = _to_money(invoice_row[8]) == _to_money(
            item_totals_by_invoice.get(invoice_id, 0)
        )
        lines_match = invoice_line_valid.get(invoice_id, False)
        invoice_reconciliation[invoice_id] = (
            "OK" if has_items and totals_match and lines_match else "CHECK"
        )

    invoice_ids_by_client: dict[int, list[int]] = {}
    for invoice_id, row in invoice_by_id.items():
        invoice_ids_by_client.setdefault(int(row[1] or 0), []).append(invoice_id)

    for key, client in clients.items():
        client_invoice_ids = invoice_ids_by_client.get(key, [])
        if client_invoice_ids:
            client["reconciliation"] = (
                "OK"
                if all(invoice_reconciliation[invoice_id] == "OK" for invoice_id in client_invoice_ids)
                else "CHECK"
            )

    sorted_clients = sorted(
        clients.values(),
        key=lambda client: _client_sort_key(
            client["client_name"], client["client_id"]
        )
    )

    client_summary_rows = [
        [
            client["client_id"],
            client["client_name"],
            client["invoice_count"],
            client["billed_quantity"],
            client["billed_amount"],
            client["paid_amount"],
            client["outstanding_amount"],
            client["payments_in_period"],
            client["trip_count"],
            client["delivered_in_period"],
            client["returned_in_period"],
            client["uninvoiced_delivered"],
            client["reconciliation"]
        ]
        for client in sorted_clients
    ]

    if client_summary_rows:
        client_summary_rows.append([
            None,
            "ALL CLIENTS",
            sum(row[2] for row in client_summary_rows),
            sum(row[3] for row in client_summary_rows),
            _to_money(sum(row[4] for row in client_summary_rows)),
            _to_money(sum(row[5] for row in client_summary_rows)),
            _to_money(sum(row[6] for row in client_summary_rows)),
            _to_money(sum(row[7] for row in client_summary_rows)),
            sum(row[8] for row in client_summary_rows),
            sum(row[9] for row in client_summary_rows),
            sum(row[10] for row in client_summary_rows),
            sum(row[11] for row in client_summary_rows),
            "OK" if all(row[12] in {"OK", "N/A"} for row in client_summary_rows) else "CHECK"
        ])

    detail_rows: list[list] = []
    invoices_by_client: dict[int, list[list]] = {}
    for row in invoice_rows:
        invoices_by_client.setdefault(int(row[1] or 0), []).append(row)

    summary_by_client = {
        int(row[0] or 0): row for row in client_summary_rows if row[0] is not None
    }

    for client in sorted_clients:
        key = int(client["client_id"] or 0)
        client_invoices = invoices_by_client.get(key, [])
        for invoice_row in client_invoices:
            invoice_id = invoice_row[0]
            items = items_by_invoice.get(invoice_id, [])
            if not items:
                detail_rows.append([
                    invoice_row[1], invoice_row[2], invoice_row[5], invoice_id,
                    invoice_row[4], invoice_row[3], None, None, 0, 0, 0,
                    invoice_row[8], invoice_row[9], invoice_row[10], invoice_row[11],
                    "CHECK - NO LINE ITEMS"
                ])
                continue

            for item_index, item_row in enumerate(items):
                line_is_valid = _to_money(item_row[5] * item_row[6]) == _to_money(item_row[7])
                invoice_is_valid = invoice_reconciliation[invoice_id] == "OK"
                validation = "OK"
                if not line_is_valid:
                    validation = "CHECK - LINE TOTAL"
                elif not invoice_is_valid:
                    validation = "CHECK - INVOICE TOTAL"

                first_line = item_index == 0
                detail_rows.append([
                    invoice_row[1],
                    (invoice_row[2] or "Unknown Client").strip(),
                    invoice_row[5],
                    invoice_id,
                    invoice_row[4],
                    invoice_row[3],
                    item_row[3],
                    item_row[4],
                    item_row[5],
                    item_row[6],
                    item_row[7],
                    invoice_row[8] if first_line else None,
                    invoice_row[9] if first_line else None,
                    invoice_row[10] if first_line else None,
                    invoice_row[11] if first_line else None,
                    validation
                ])

        if client_invoices:
            summary_row = summary_by_client[key]
            detail_rows.append([
                client["client_id"],
                f'{client["client_name"]} TOTAL',
                None, None, None, None, None, None,
                summary_row[3],
                None,
                summary_row[4],
                summary_row[4],
                summary_row[5],
                summary_row[6],
                sum(int(row[11] or 0) for row in client_invoices),
                summary_row[12]
            ])

    totals = {
        "billed_quantity": sum(row[3] for row in client_summary_rows if row[1] != "ALL CLIENTS"),
        "uninvoiced_delivered": sum(
            row[11] for row in client_summary_rows if row[1] != "ALL CLIENTS"
        ),
        "reconciliation": (
            "OK"
            if all(row[12] in {"OK", "N/A"} for row in client_summary_rows)
            else "CHECK"
        )
    }
    return client_summary_rows, detail_rows, totals


def build_excel_export(
    db: Session,
    report_type: ExportReportType = "full",
    from_date: date | None = None,
    to_date: date | None = None,
    client_id: int | None = None,
    driver_id: int | None = None,
    status: str | None = None,
    payment_method: str | None = None,
    include_cancelled: bool = False
) -> tuple[bytes, str, dict]:
    normalized_status = (status or "").strip().lower() or None
    if normalized_status and normalized_status not in ALLOWED_STATUSES:
        raise ValueError("Invalid status filter")

    start_dt, end_dt = _to_datetime_range(from_date, to_date)

    invoices, invoice_rows, invoice_item_rows = _build_invoices_data(
        db=db,
        start_dt=start_dt,
        end_dt=end_dt,
        client_id=client_id,
        driver_id=driver_id,
        status=normalized_status,
        include_cancelled=include_cancelled
    )

    trip_rows = _build_trips_data(
        db=db,
        start_dt=start_dt,
        end_dt=end_dt,
        client_id=client_id,
        driver_id=driver_id,
        status=normalized_status,
        include_cancelled=include_cancelled
    )

    payment_rows = _build_payments_data(
        db=db,
        start_dt=start_dt,
        end_dt=end_dt,
        client_id=client_id,
        driver_id=driver_id,
        status=normalized_status,
        payment_method=payment_method,
        include_cancelled=include_cancelled
    )

    client_summary_rows, client_detail_rows, client_totals = _build_client_views(
        invoice_rows=invoice_rows,
        invoice_item_rows=invoice_item_rows,
        trip_rows=trip_rows,
        payment_rows=payment_rows
    )

    total_invoiced_amount = _to_money(sum(row[8] for row in invoice_rows))
    total_paid_amount = _to_money(sum(row[9] for row in invoice_rows))
    total_outstanding_amount = _to_money(sum(row[10] for row in invoice_rows))
    total_payment_amount = _to_money(sum(row[7] for row in payment_rows))
    total_delivered_jars = int(sum(row[8] for row in trip_rows))
    total_returned_jars = int(sum(row[9] for row in trip_rows))

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    summary_headers = ["Metric", "Value"]
    summary_rows = [
        ["Generated At (IST)", _to_ist_datetime(datetime.utcnow())],
        ["Report Type", report_type],
        ["From Date", from_date.isoformat() if from_date else "All"],
        ["To Date", to_date.isoformat() if to_date else "All"],
        ["Timezone", "Asia/Kolkata (IST)"],
        [
            "Date Filter Rules",
            "Invoices by invoice date; trips by delivery date; payments by payment date"
        ],
        ["Client ID", client_id if client_id else "All"],
        ["Driver ID", driver_id if driver_id else "All"],
        ["Invoice Status Filter", normalized_status or "All"],
        ["Payment Method Filter", payment_method or "All"],
        ["Include Cancelled", "Yes" if include_cancelled else "No"],
        ["Invoices Count", len(invoices)],
        ["Invoice Items Count", len(invoice_item_rows)],
        ["Trips Count", len(trip_rows)],
        ["Payments Count", len(payment_rows)],
        ["Total Invoiced Amount", total_invoiced_amount],
        ["Total Invoice Paid Amount", total_paid_amount],
        ["Total Invoice Outstanding", total_outstanding_amount],
        ["Total Payments Amount", total_payment_amount],
        ["Total Billed Container Quantity", client_totals["billed_quantity"]],
        ["Total Delivered During Period", total_delivered_jars],
        ["Total Returned During Period", total_returned_jars],
        ["Total Uninvoiced Delivered During Period", client_totals["uninvoiced_delivered"]],
        ["Invoice Reconciliation", client_totals["reconciliation"]],
    ]
    _write_sheet(summary_ws, summary_headers, summary_rows)

    client_summary_ws = wb.create_sheet("Client_Summary")
    _write_sheet(
        client_summary_ws,
        [
            "Client ID",
            "Client Name",
            "Invoice Count",
            "Billed Container Quantity",
            "Billed Amount",
            "Invoice Paid Amount",
            "Outstanding Amount",
            "Payments Amount Recorded During Period",
            "Trips During Period",
            "Delivered During Period",
            "Returned During Period",
            "Uninvoiced Delivered During Period",
            "Reconciliation"
        ],
        client_summary_rows
    )
    client_summary_ws.freeze_panes = "C2"

    if report_type in {"invoices", "full"}:
        client_details_ws = wb.create_sheet("Client_Details")
        _write_sheet(
            client_details_ws,
            [
                "Client ID",
                "Client Name",
                "Invoice Date (IST)",
                "Invoice ID",
                "Status",
                "Driver Name(s)",
                "Container ID",
                "Container Name",
                "Quantity",
                "Unit Price",
                "Line Total",
                "Invoice Total",
                "Amount Paid",
                "Outstanding Amount",
                "Trip Count",
                "Validation"
            ],
            client_detail_rows
        )
        client_details_ws.freeze_panes = "C2"

    if report_type in {"invoices", "full"}:
        invoices_ws = wb.create_sheet("Invoices")
        _write_sheet(
            invoices_ws,
            [
                "Invoice ID",
                "Client ID",
                "Client Name",
                "Driver Name(s)",
                "Status",
                "Created At (IST)",
                "Due Date (IST)",
                "Confirmed At (IST)",
                "Total Amount",
                "Amount Paid",
                "Outstanding Amount",
                "Trip Count"
            ],
            invoice_rows
        )

        invoice_items_ws = wb.create_sheet("Invoice_Items")
        _write_sheet(
            invoice_items_ws,
            [
                "Invoice ID",
                "Client ID",
                "Client Name",
                "Container ID",
                "Container Name",
                "Quantity",
                "Price Snapshot",
                "Line Total"
            ],
            invoice_item_rows
        )

    if report_type in {"billing", "full"}:
        trips_ws = wb.create_sheet("Bills_Trips")
        _write_sheet(
            trips_ws,
            [
                "Trip ID",
                "Trip DateTime (IST)",
                "Client ID",
                "Client Name",
                "Driver ID",
                "Driver Name",
                "Invoice ID",
                "Invoice Status",
                "Delivered Jars",
                "Returned Jars",
                "Container Breakdown"
            ],
            trip_rows
        )

    if report_type in {"payments", "full"}:
        payments_ws = wb.create_sheet("Payments")
        _write_sheet(
            payments_ws,
            [
                "Payment ID",
                "Payment DateTime (IST)",
                "Invoice ID",
                "Client ID",
                "Client Name",
                "Invoice Status",
                "Payment Method",
                "Amount",
                "Cash Amount",
                "UPI Amount",
                "UPI Account"
            ],
            payment_rows
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    exported_at = datetime.now(IST).strftime("%Y%m%d_%H%M%S")
    filename = f"rivarich_{report_type}_export_{exported_at}.xlsx"

    export_meta = {
        "invoices_count": len(invoices),
        "invoice_items_count": len(invoice_item_rows),
        "trips_count": len(trip_rows),
        "payments_count": len(payment_rows),
        "total_invoiced": total_invoiced_amount,
        "total_outstanding": total_outstanding_amount,
        "total_payments": total_payment_amount,
        "reconciliation": client_totals["reconciliation"]
    }

    return output.getvalue(), filename, export_meta
