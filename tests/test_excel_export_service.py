from datetime import datetime
from io import BytesIO
import unittest
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app.services.excel_export_service import (
    _build_client_views,
    _to_ist_datetime,
    _write_sheet,
    build_excel_export,
)


class ExcelExportServiceTests(unittest.TestCase):
    def setUp(self):
        self.invoice_rows = [
            [1, 7, "Axis Bank", "Driver 1", "pending", datetime(2026, 6, 2, 10), None, None, 170, 50, 120, 2],
            [2, 7, "Axis Bank", "Driver 1", "pending", datetime(2026, 6, 3, 10), None, None, 105, 0, 105, 1],
            [3, 8, "Beta Client", "Driver 2", "paid", datetime(2026, 6, 4, 10), None, None, 40, 40, 0, 1],
        ]
        self.item_rows = [
            [1, 7, "Axis Bank", 2, "20 Ltr Jar", 2, 35, 70],
            [1, 7, "Axis Bank", 3, "Bottle Case", 1, 100, 100],
            [2, 7, "Axis Bank", 2, "20 Ltr Jar", 3, 35, 105],
            [3, 8, "Beta Client", 2, "20 Ltr Jar", 1, 40, 40],
        ]
        self.trip_rows = [
            [10, datetime(2026, 6, 2, 9), 7, "Axis Bank", 1, "Driver 1", 1, "pending", 3, 3, "20 Ltr Jar"],
            [11, datetime(2026, 6, 3, 9), 7, "Axis Bank", 1, "Driver 1", None, None, 2, 0, "20 Ltr Jar"],
            [12, datetime(2026, 6, 4, 9), 8, "Beta Client", 2, "Driver 2", 3, "paid", 1, 1, "20 Ltr Jar"],
        ]
        self.payment_rows = [
            [20, datetime(2026, 6, 5, 9), 1, 7, "Axis Bank", "pending", "CASH", 50, 50, 0, None]
        ]

    def test_client_summary_groups_and_reconciles(self):
        summary_rows, detail_rows, totals = _build_client_views(
            self.invoice_rows,
            self.item_rows,
            self.trip_rows,
            self.payment_rows,
        )

        axis = summary_rows[0]
        self.assertEqual(axis[:4], [7, "Axis Bank", 2, 6])
        self.assertEqual(axis[4:8], [275.0, 50.0, 225.0, 50.0])
        self.assertEqual(axis[8:13], [2, 5, 3, 2, "OK"])
        self.assertEqual(summary_rows[-1][1], "ALL CLIENTS")
        self.assertEqual(totals, {
            "billed_quantity": 7,
            "uninvoiced_delivered": 2,
            "reconciliation": "OK",
        })

        axis_detail_names = [row[1] for row in detail_rows[:4]]
        self.assertEqual(axis_detail_names, [
            "Axis Bank", "Axis Bank", "Axis Bank", "Axis Bank TOTAL"
        ])
        self.assertEqual(detail_rows[0][11:15], [170, 50, 120, 2])
        self.assertEqual(detail_rows[1][11:15], [None, None, None, None])

    def test_reconciliation_flags_bad_line_total(self):
        bad_items = [row.copy() for row in self.item_rows]
        bad_items[0][7] = 69

        summary_rows, detail_rows, totals = _build_client_views(
            self.invoice_rows,
            bad_items,
            self.trip_rows,
            self.payment_rows,
        )

        self.assertEqual(summary_rows[0][12], "CHECK")
        self.assertEqual(detail_rows[0][15], "CHECK - LINE TOTAL")
        self.assertEqual(totals["reconciliation"], "CHECK")

    def test_utc_timestamp_is_exported_as_ist(self):
        converted = _to_ist_datetime(datetime(2026, 6, 1, 0, 0))
        self.assertEqual(converted, datetime(2026, 6, 1, 5, 30))

    def test_sheet_format_survives_xlsx_round_trip(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Client_Summary"
        _write_sheet(
            sheet,
            ["Client Name", "Billed Amount", "Delivered During Period"],
            [["Axis Bank", 1400, 40], ["ALL CLIENTS", 1400, 40]],
        )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        loaded = load_workbook(output, data_only=True)
        loaded_sheet = loaded["Client_Summary"]

        self.assertEqual(loaded_sheet["A2"].value, "Axis Bank")
        self.assertEqual(loaded_sheet["B2"].number_format, "₹#,##0.00")
        self.assertEqual(loaded_sheet["C2"].number_format, "#,##0")
        self.assertEqual(loaded_sheet.freeze_panes, "A2")
        self.assertFalse(loaded_sheet.sheet_view.showGridLines)

    def test_full_export_places_client_views_first(self):
        with (
            patch(
                "app.services.excel_export_service._build_invoices_data",
                return_value=([object(), object(), object()], self.invoice_rows, self.item_rows),
            ),
            patch(
                "app.services.excel_export_service._build_trips_data",
                return_value=self.trip_rows,
            ),
            patch(
                "app.services.excel_export_service._build_payments_data",
                return_value=self.payment_rows,
            ),
        ):
            file_bytes, filename, meta = build_excel_export(
                db=object(),
                report_type="full",
            )

        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        self.assertEqual(workbook.sheetnames[:3], [
            "Summary", "Client_Summary", "Client_Details"
        ])
        self.assertEqual(workbook["Client_Summary"]["B2"].value, "Axis Bank")
        self.assertEqual(workbook["Client_Summary"]["E2"].value, 275)
        self.assertEqual(workbook["Client_Details"]["B2"].value, "Axis Bank")
        self.assertEqual(workbook["Invoices"]["F1"].value, "Created At (IST)")
        self.assertEqual(meta["reconciliation"], "OK")
        self.assertTrue(filename.startswith("rivarich_full_export_"))


if __name__ == "__main__":
    unittest.main()
